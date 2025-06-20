import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Set page configuration
st.set_page_config(page_title="PREDICTIVE SUMMARIZER", page_icon="📊")

# Title and description
st.title("PREDICTIVE SUMMARIZER")
st.markdown("NEKENNAV")

# File uploader widget
uploaded_files = st.file_uploader(
    "",  # Removed "Choose Excel files to merge"
    accept_multiple_files=True,
    type=['xlsx', 'xls']
)

# Create a directory to store uploaded files temporarily
UPLOAD_DIR = "uploads"
if not os.path.exists(UPLOAD_DIR):
    os.makedirs(UPLOAD_DIR)

# Function to convert time to seconds
def time_to_seconds(time_val):
    try:
        if pd.isna(time_val):
            return 0
        if isinstance(time_val, (int, float)):
            return float(time_val)
        if isinstance(time_val, str):
            parts = time_val.split(':')
            parts = [p.strip() for p in parts]
            if len(parts) == 3:  # HH:MM:SS
                h, m, s = map(int, parts)
                return h * 3600 + m * 60 + s
            elif len(parts) == 2:  # MM:SS
                m, s = map(int, parts)
                return m * 60 + s
        return 0
    except:
        return 0

# Function to format seconds to [h]:mm:ss for display
def seconds_to_time(seconds):
    if pd.isna(seconds):
        return "0:00:00"
    seconds = int(seconds)
    hours = seconds // 3600
    seconds %= 3600
    minutes = seconds // 60
    seconds = seconds % 60
    return f"{hours}:{minutes:02d}:{seconds:02d}"

# Function to merge and aggregate Excel files
def merge_excel_files(files):
    try:
        # List to store DataFrames
        dfs = []
        
        # Read each Excel file
        for file in files:
            # Save the file temporarily
            file_path = os.path.join(UPLOAD_DIR, file.name)
            with open(file_path, "wb") as f:
                f.write(file.getbuffer())
            
            # Read the first sheet of the Excel file
            df = pd.read_excel(file_path, engine=None)
            dfs.append(df)
        
        # Concatenate all DataFrames
        if dfs:
            merged_df = pd.concat(dfs, ignore_index=True)
            
            # Remove SNo., Total Calls, and Pause Count columns if they exist
            columns_to_drop = ['SNo.', 'Total Calls', 'Pause Count']
            merged_df = merged_df.drop(columns=[col for col in columns_to_drop if col in merged_df.columns])
                
            # Remove rows where Collector Name is blank (NaN, empty string, or whitespace)
            if 'Collector Name' in merged_df.columns:
                merged_df = merged_df[merged_df['Collector Name'].notna() & (merged_df['Collector Name'].str.strip() != '')]
            else:
                return None, "Collector Name column not found in the data."
            
            # Define time columns to convert and sum
            time_columns = [
                'Spent Time', 'Talk Time', 'AVG Talk Time', 'Wait Time',
                'Average Wait Time', 'Write Time', 'AVG Write Time', 'Pause Time'
            ]
            
            # Filter time columns to those present in the DataFrame
            valid_time_columns = [col for col in time_columns if col in merged_df.columns]
            
            # Convert time columns to seconds for aggregation
            for col in valid_time_columns:
                merged_df[col] = merged_df[col].apply(time_to_seconds)
            
            # Group by Collector Name and aggregate
            agg_dict = {}
            for col in valid_time_columns:
                agg_dict[col] = 'sum'
            other_columns = [col for col in merged_df.columns if col not in valid_time_columns + ['Collector Name']]
            for col in other_columns:
                agg_dict[col] = 'first'
            
            if not agg_dict:
                return merged_df, None
            
            merged_df = merged_df.groupby('Collector Name').agg(agg_dict).reset_index()
            
            # Calculate averages for time columns
            avg_row = {'Collector Name': 'Average'}
            for col in valid_time_columns:
                avg_row[col] = merged_df[col].mean()
            for col in other_columns:
                avg_row[col] = None
            
            # Append average row to DataFrame
            avg_df = pd.DataFrame([avg_row])
            merged_df = pd.concat([merged_df, avg_df], ignore_index=True)
            
            return merged_df, None
        else:
            return None, "No valid Excel files uploaded."
            
    except Exception as e:
        return None, f"Error merging files: {str(e)}"

# Process uploaded files
if uploaded_files:
    st.success(f"Successfully uploaded {len(uploaded_files)} file(s)!")
    
    # Merge files
    merged_df, error = merge_excel_files(uploaded_files)
    
    if error:
        st.error(error)
    else:
        # Display preview of merged data
        st.write("**Preview of Merged Data**")
        display_df = merged_df.copy()
        time_columns = [
            'Spent Time', 'Talk Time', 'AVG Talk Time', 'Wait Time',
            'Average Wait Time', 'Write Time', 'AVG Write Time', 'Pause Time'
        ]
        valid_time_columns = [col for col in time_columns if col in display_df.columns]
        for col in valid_time_columns:
            display_df[col] = display_df[col].apply(seconds_to_time)
        
        preview_text = display_df.to_string(index=False)
        st.text_area(
            "Merged Data Preview",
            preview_text,
            height=200,
            key=f"preview_merged_{hash(str(uploaded_files))}"
        )
        
        # Prepare download for merged data
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            excel_df = merged_df.copy()
            for col in valid_time_columns:
                excel_df[col] = excel_df[col] / 86400.0
            
            excel_df.to_excel(writer, index=False, sheet_name='Sheet1')
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            
            for col in valid_time_columns:
                col_idx = merged_df.columns.get_loc(col) + 1
                col_letter = get_column_letter(col_idx)
                for row in range(2, len(merged_df) + 2):
                    cell = worksheet[f"{col_letter}{row}"]
                    cell.number_format = '[h]:mm:ss'
                    cell.alignment = Alignment(horizontal='right')
                
                header_cell = worksheet[f"{col_letter}1"]
                header_cell.alignment = Alignment(horizontal='right')
            
            for col_idx in range(1, len(merged_df.columns) + 1):
                col_name = merged_df.columns[col_idx - 1]
                if col_name != 'Collector Name':
                    col_letter = get_column_letter(col_idx)
                    for row in range(1, len(merged_df) + 2):
                        cell = worksheet[f"{col_letter}{row}"]
                        if col_name not in valid_time_columns:
                            cell.alignment = Alignment(horizontal='right')
        
        output.seek(0)
        
        st.download_button(
            label="Download Merged Excel File",
            data=output,
            file_name=f"Merged_Excel_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_merged"
        )

else:
    st.info("Please upload one or more Excel files to merge.")
